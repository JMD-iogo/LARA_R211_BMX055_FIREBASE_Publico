import os
import openpyxl
import serial  # pip install pyserttial
import time
import logging

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")

def capturar_dados_serial(porta, baudrate, nome_arquivo, tempo_aquisicao):
    """
    Captura dados da porta serial e os armazena em um arquivo Excel,
    criando uma nova aba com nome fornecido pelo usuário.
    
    Parâmetros:
    - porta: porta serial (ex. "COM10").
    - baudrate: taxa de transmissão.
    - nome_arquivo: nome do arquivo Excel a ser salvo.
    - tempo_aquisicao: tempo de aquisição em segundos.
    """
    try:
        # Solicitar o nome da aba ao usuário
        nome_pagina = input(f"Escreva o nome da folha nova: '{nome_arquivo}': ")

        # Verifica se o arquivo já existe
        if os.path.exists(nome_arquivo):
            # Carregar o arquivo existente
            workbook = openpyxl.load_workbook(nome_arquivo)
            # Criar uma nova aba com o nome fornecido pelo usuário
            sheet = workbook.create_sheet(title=nome_pagina)
            logging.info(f"Folha '{nome_pagina}' criada")
        else:
            # Caso o arquivo não exista, cria um novo arquivo com a aba especificada
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = nome_pagina
            logging.info(f"Novo excel e folha '{nome_pagina}' criados.")

        n = 0
        tempo_marca = time.time()
        tempo_contar = 0
        
        with serial.Serial(porta, baudrate, timeout=1) as ser:
            tempo_inicio = time.time()
            tempo_valido = 0  # Tempo efetivo de aquisição válido
            nomes_cabecalho = []
            linha_atual = 2
            ultimo_tempo_valido = time.time()

            while tempo_valido < tempo_aquisicao:
                linha = ser.readline().decode().strip()
                if linha:
                    partes = linha.split(',')
                   
                    # Verifica se a quantidade de valores é par (nome, valor)
                    if len(partes) % 2 == 0:
                        n = n + 1
                        if n == 1:
                            tempo_contar = time.time() - tempo_marca  # Elimina possíveis tempos passados antes de adquirir dados certos
                        tempo_atual = time.time()
                        tempo_valido += tempo_atual - ultimo_tempo_valido - tempo_contar
                        ultimo_tempo_valido = tempo_atual
                        tempo_contar = 0

                        print(f"Tempo decorrido: {tempo_valido:.2f} segundos (de{tempo_aquisicao})")  # Exibe tempo na console

                        # Extrair nomes e valores
                        nomes = partes[::2]
                        valores = partes[1::2]

                        # Definir cabeçalho apenas uma vez
                        if not nomes_cabecalho:
                            nomes_cabecalho.extend(nomes)
                            for i, nome in enumerate(nomes_cabecalho, start=1):
                                sheet.cell(row=1, column=i, value=nome)

                        # Adicionar os valores na linha correta
                        for i, valor in enumerate(valores, start=1):
                            sheet.cell(row=linha_atual, column=i, value=valor)
                        
                        # Exibir na consola os dados que estão sendo salvos
                        print(dict(zip(nomes, valores)))

                        linha_atual += 1
                    else:
                        # Se os dados forem inválidos, apenas atualizamos a referência de tempo
                        logging.warning(f"Dados inválidos recebidos: {linha}")
        
        # Salvar o arquivo Excel
        workbook.save(nome_arquivo)
        print(f"Arquivo salvo como {nome_arquivo}")
    
    except serial.SerialException as e:
        logging.error(f"Erro ao acessar a porta serial: {e}")

# Funcao ira ler a porta especificada no rate especificado, e adquire dados no tempo especificado, 
#cria um doc exel novo caso nao exista nenhum com o nome e pede para 
#caso exista ele simplesmente pede o nome da folha, caso introduza um nome ja colocado ele coloca um 1 a frente (e por ai diante)
#os dados tem de vir na forma Serial.print("dado a,");Serial.print(a);Serial.print(",dado b,");Serial.print(b)...
capturar_dados_serial("COM10", 115200, "dados.xlxs", 15)
